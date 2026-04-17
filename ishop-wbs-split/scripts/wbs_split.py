#!/usr/bin/env python3
"""
智能网页需求WBS拆分工具

通过 Qoder AI 辅助从网页内容提取表格数据并进行WBS拆分
支持蓝湖、Jira、Confluence等包含表格的网页

使用示例:
    python wbs_split.py --url "网页链接"
    python wbs_split.py --url "网页链接" --output "需求_WBS.xlsx"
    python wbs_split.py --url "网页链接" --keep-modules "模块1,模块2" --demand-map "模块1:需求1"

注意:
    本工具需要与 Qoder AI 配合使用
    对于蓝湖页面，AI 会自动调用 lanhu MCP 工具获取数据
    对于其他网页，AI 会使用 fetch_content 工具读取内容
"""

import argparse
import sys
import json
import re
from typing import List, Dict, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


# 默认排除的模块
DEFAULT_EXCLUDE_MODULES = ['首页模块', '自定义首页模块', '我的页面模块', '商品档案']


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='智能网页需求WBS拆分工具 - 通过 Qoder AI 辅助从网页提取表格数据并拆分需求'
    )
    parser.add_argument(
        '--url',
        required=True,
        help='需求页面URL（支持蓝湖、Jira、Confluence等任意包含表格的网页）'
    )
    parser.add_argument(
        '--data-file',
        help='可选：直接提供包含表格数据的JSON文件路径（跳过网页抓取）'
    )
    parser.add_argument(
        '--output',
        default='./需求_WBS.xlsx',
        help='输出Excel文件路径（默认: ./需求_WBS.xlsx）'
    )
    parser.add_argument(
        '--target-platform',
        default='app',
        choices=['app', 'pc', 'all'],
        help='目标平台：app（仅APP端）、pc（仅PC端）、all（全部平台）。默认：app'
    )
    parser.add_argument(
        '--max-task-hours',
        type=float,
        default=3.0,
        help='单个任务的最大工时（小时）。默认：3小时'
    )
    parser.add_argument(
        '--keep-modules',
        help='保留的模块列表，用逗号分隔'
    )
    parser.add_argument(
        '--exclude-modules',
        help='排除的模块列表，用逗号分隔'
    )
    parser.add_argument(
        '--demand-map',
        help='模块到需求名称的映射，格式：模块1:需求1,模块2:需求2'
    )
    return parser.parse_args()


def parse_modules(modules_str):
    """解析模块列表"""
    if not modules_str:
        return []
    return [m.strip() for m in modules_str.split(',')]


def parse_demand_map(map_str):
    """解析需求映射"""
    if not map_str:
        return {}
    result = {}
    for item in map_str.split(','):
        if ':' in item:
            module, demand = item.split(':', 1)
            result[module.strip()] = demand.strip()
    return result


def is_lanhu_url(url: str) -> bool:
    """判断是否为蓝湖URL"""
    return 'lanhuapp.com' in url


def print_guide_message(url: str) -> None:
    """
    打印操作指引信息
    
    Args:
        url: 网页URL
    """
    print("\n" + "="*70)
    print("📋 WBS拆分工具 - 数据获取指引")
    print("="*70)
    
    if is_lanhu_url(url):
        print("\n🔥 检测到蓝湖页面")
        print("\n请按以下步骤操作：")
        print("\n1️⃣  请告诉 AI：")
        print("   \"请使用 lanhu_get_ai_analyze_page_result 工具读取这个蓝湖页面的表格数据\"")
        print("\n2️⃣  AI 获取数据后，会自动将数据保存为 JSON 文件")
        print("\n3️⃣  然后重新运行此工具，添加 --data-file 参数：")
        print(f"   python wbs_split.py --data-file <数据文件.json> --output <输出.xlsx>")
        print("\n💡 提示：AI 已经可以访问蓝湖内容，请直接向 AI 发送上述请求")
    else:
        print("\n🌐 检测到通用网页")
        print("\n请按以下步骤操作：")
        print("\n1️⃣  请告诉 AI：")
        print("   \"请使用 fetch_content 工具读取这个网页的表格数据\"")
        print("\n2️⃣  AI 获取数据后，请要求 AI：")
        print("   \"请将表格数据提取为 JSON 格式并保存\"")
        print("\n3️⃣  然后重新运行此工具，添加 --data-file 参数：")
        print(f"   python wbs_split.py --data-file <数据文件.json> --output <输出.xlsx>")
    
    print("\n" + "="*70)
    print("\n⏸️  工具已暂停，等待数据文件...")
    print("\n")


def load_data_from_json(json_file: str) -> tuple:
    """
    从JSON文件加载表格数据
    
    Args:
        json_file: JSON文件路径
    
    Returns:
        tuple: (headers, data) 表头和数据列表
    
    Expected JSON format:
        {
            "headers": ["需求名称", "模块", "任务", ...],
            "data": [
                {"需求名称": "xxx", "模块": "xxx", "任务": "xxx", ...},
                ...
            ]
        }
    """
    print(f"\n📂 从文件加载数据: {json_file}")
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # 检查数据格式
        if 'headers' not in json_data or 'data' not in json_data:
            raise ValueError("JSON文件格式错误，需要包含 'headers' 和 'data' 字段")
        
        headers = json_data['headers']
        data = json_data['data']
        
        print(f"✓ 成功加载 {len(data)} 行数据")
        print(f"✓ 表头: {headers}")
        
        return headers, data
        
    except FileNotFoundError:
        print(f"\n❌ 错误：找不到文件 {json_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"\n❌ 错误：JSON文件格式错误 - {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 错误：加载数据失败 - {e}")
        sys.exit(1)




def parse_time_to_hours(time_str: str) -> float:
    """
    解析时间字符串为小时数
    
    支持格式：
    - "2h" -> 2.0
    - "3小时" -> 3.0
    - "1.5h" -> 1.5
    - "2天" -> 16.0 (按8小时/天)
    - "1d" -> 8.0
    
    Args:
        time_str: 时间字符串
    
    Returns:
        float: 小时数，无法解析返回 0
    """
    if not time_str or not isinstance(time_str, str):
        return 0.0
    
    time_str = time_str.strip().lower()
    
    # 匹配数字
    match = re.match(r'([\d.]+)\s*(h|小时|d|天|day)?', time_str)
    if not match:
        return 0.0
    
    value = float(match.group(1))
    unit = match.group(2) or 'h'
    
    # 转换为小时
    if unit in ['d', '天', 'day']:
        return value * 8  # 1天 = 8小时
    else:
        return value


def validate_task_granularity(data: List[Dict], max_hours: float) -> tuple:
    """
    验证任务粒度是否符合要求
    
    Args:
        data: 任务数据列表
        max_hours: 最大工时限制
    
    Returns:
        tuple: (是否通过, 违规任务列表)
    """
    violations = []
    
    time_key = '时间' if any('时间' in row for row in data) else 'time'
    task_key = '任务' if any('任务' in row for row in data) else 'task'
    
    for idx, row in enumerate(data, 1):
        time_str = row.get(time_key, '')
        task_name = row.get(task_key, '')
        
        if time_str:
            hours = parse_time_to_hours(time_str)
            if hours > max_hours:
                violations.append({
                    'row': idx,
                    'task': task_name,
                    'time': time_str,
                    'hours': hours,
                    'max_hours': max_hours
                })
    
    return len(violations) == 0, violations


def filter_by_platform(data: List[Dict], target_platform: str) -> List[Dict]:
    """
    根据目标平台过滤任务
    
    Args:
        data: 任务数据列表
        target_platform: 目标平台 (app/pc/all)
    
    Returns:
        list: 过滤后的数据
    """
    if target_platform == 'all':
        return data
    
    # 识别列名
    module_key = '模块' if any('模块' in row for row in data) else 'module'
    task_key = '任务' if any('任务' in row for row in data) else 'task'
    
    # 平台关键词
    platform_keywords = {
        'app': ['app', 'APP', '移动端', '手机', 'mobile', '首页', '我的页面', '消息中心'],
        'pc': ['pc', 'PC', 'web', 'Web', '电脑端', '管理后台', '后台管理']
    }
    
    keywords = platform_keywords.get(target_platform, [])
    
    filtered = []
    for row in data:
        module = row.get(module_key, '')
        task = row.get(task_key, '')
        combined_text = f"{module} {task}".lower()
        
        # 检查是否包含目标平台关键词
        if target_platform == 'app':
            # APP端：排除明确标记为PC的内容
            is_pc = any(kw.lower() in combined_text for kw in platform_keywords['pc'])
            if not is_pc:
                filtered.append(row)
        elif target_platform == 'pc':
            # PC端：只保留明确标记为PC的内容
            is_pc = any(kw.lower() in combined_text for kw in platform_keywords['pc'])
            if is_pc:
                filtered.append(row)
    
    return filtered


def filter_and_process_data(headers, data, keep_modules, exclude_modules, demand_map):
    """
    过滤和处理数据
    
    Args:
        headers: 表头列表
        data: 数据字典列表
        keep_modules: 保留的模块列表
        exclude_modules: 排除的模块列表
        demand_map: 模块到需求名称的映射
    
    Returns:
        list: 处理后的数据列表
    """
    print(f"\n🔧 处理数据...")
    
    # 确定列名（支持中英文）
    module_key = '模块' if '模块' in headers else 'module'
    task_key = '任务' if '任务' in headers else 'task'
    demand_key = '需求名称' if '需求名称' in headers else 'demand'
    
    filtered_data = []
    for row in data:
        module = row.get(module_key, '')
        task = row.get(task_key, '')
        
        # 过滤空任务
        if not task:
            continue
        
        # 排除模块
        if module and module in exclude_modules:
            continue
        
        # 保留模块过滤
        if keep_modules and module not in keep_modules:
            continue
        
        # 填充需求名称
        if module and module in demand_map:
            row[demand_key] = demand_map[module]
        
        filtered_data.append(row)
    
    print(f"✓ 过滤后保留 {len(filtered_data)} 行有效数据")
    
    return filtered_data


def export_to_excel(headers, data, output_path):
    """
    导出数据到Excel
    
    Args:
        headers: 表头列表
        data: 数据字典列表
        output_path: 输出文件路径
    """
    print(f"\n📝 导出到 Excel: {output_path}")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WBS"
    
    # 写入表头
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(header, '')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    wb.save(output_path)
    print(f"✓ 成功导出 {len(data)} 行数据")


def display_summary(data, headers):
    """显示WBS拆分摘要"""
    print("\n" + "="*60)
    print("📋 WBS拆分结果摘要")
    print("="*60)
    
    demand_key = '需求名称' if '需求名称' in headers else 'demand'
    module_key = '模块' if '模块' in headers else 'module'
    task_key = '任务' if '任务' in headers else 'task'
    
    current_demand = None
    current_module = None
    task_count = 0
    
    for row in data:
        demand = row.get(demand_key, '')
        module = row.get(module_key, '')
        task = row.get(task_key, '')
        
        if demand != current_demand:
            if current_demand:
                print(f"    └─ 共 {task_count} 个任务")
            print(f"\n📌 {demand or '未分类'}")
            current_demand = demand
            current_module = None
            task_count = 0
        
        if module != current_module:
            if current_module:
                print(f"    └─ {task_count} 个任务")
                task_count = 0
            print(f"  📂 {module}")
            current_module = module
        
        print(f"    ✓ {task}")
        task_count += 1
    
    if current_module:
        print(f"    └─ {task_count} 个任务")
    
    print("\n" + "="*60)
    print(f"📊 总计: {len(data)} 个任务")
    print("="*60)


def main_process(args):
    """主处理函数"""
    # 解析参数
    keep_modules = parse_modules(args.keep_modules)
    exclude_modules = parse_modules(args.exclude_modules) or DEFAULT_EXCLUDE_MODULES
    demand_map = parse_demand_map(args.demand_map)
    
    # 获取目标平台和最大工时
    target_platform = args.target_platform
    max_task_hours = args.max_task_hours
    
    # 显示约束信息
    print("\n" + "="*70)
    print("📋 WBS 拆分约束")
    print("="*70)
    print(f"🎯 目标平台: {target_platform.upper()}")
    print(f"⏱️  任务粒度: 最大 {max_task_hours} 小时")
    print("="*70)
    
    # 获取数据
    if args.data_file:
        # 从JSON文件加载数据
        headers, raw_data = load_data_from_json(args.data_file)
    else:
        # 没有数据文件，打印操作指引
        print_guide_message(args.url)
        print("\n💡 提示：")
        print("   1. 请按照上述指引让 AI 获取网页数据")
        print("   2. 获取数据后，使用 --data-file 参数重新运行此工具")
        print("\n示例命令：")
        print(f"   python wbs_split.py \\")
        print(f"       --url \"{args.url}\" \\")
        print(f"       --data-file table_data.json \\")
        print(f"       --target-platform {target_platform} \\")
        print(f"       --max-task-hours {max_task_hours} \\")
        print(f"       --output {args.output}")
        if args.keep_modules:
            print(f"       --keep-modules \"{args.keep_modules}\"")
        if args.demand_map:
            print(f"       --demand-map \"{args.demand_map}\"")
        sys.exit(0)
    
    # 1. 平台过滤
    if target_platform != 'all':
        print(f"\n🔍 应用平台过滤: {target_platform.upper()}")
        platform_filtered_data = filter_by_platform(raw_data, target_platform)
        print(f"✓ 平台过滤后保留 {len(platform_filtered_data)} 行数据")
    else:
        platform_filtered_data = raw_data
    
    # 2. 过滤和处理数据
    filtered_data = filter_and_process_data(
        headers=headers,
        data=platform_filtered_data,
        keep_modules=keep_modules,
        exclude_modules=exclude_modules,
        demand_map=demand_map
    )
    
    # 3. 验证任务粒度
    print(f"\n⚡ 验证任务粒度（最大 {max_task_hours} 小时）...")
    is_valid, violations = validate_task_granularity(filtered_data, max_task_hours)
    
    if not is_valid:
        print(f"\n❌ 发现 {len(violations)} 个任务超过粒度限制：\n")
        for v in violations[:10]:  # 只显示前10个
            print(f"   ⚠️  第 {v['row']} 行: {v['task']}")
            print(f"      时间: {v['time']} ({v['hours']}h) > {v['max_hours']}h")
        
        if len(violations) > 10:
            print(f"\n   ... 还有 {len(violations) - 10} 个任务超限")
        
        print(f"\n💡 建议：请将这些任务拆分为更小的子任务，每个不超过 {max_task_hours} 小时")
        print(f"\n⚠️  警告：已生成 Excel 文件，但请注意修改超限任务！")
    else:
        print(f"✓ 所有任务粒度符合要求（≤ {max_task_hours}h）")
    
    # 导出到Excel
    export_to_excel(headers, filtered_data, args.output)
    
    # 显示摘要
    display_summary(filtered_data, headers)
    
    # 显示最终统计
    print(f"\n" + "="*70)
    print(f"📊 最终统计")
    print(f"="*70)
    print(f"🎯 目标平台: {target_platform.upper()}")
    print(f"📝 任务总数: {len(filtered_data)} 个")
    
    # 计算总工时
    time_key = '时间' if '时间' in headers else 'time'
    total_hours = sum(parse_time_to_hours(row.get(time_key, '')) for row in filtered_data)
    print(f"⏱️  总工时: {total_hours:.1f} 小时")
    
    if not is_valid:
        print(f"⚠️  粒度检查: 失败（{len(violations)} 个任务超限）")
    else:
        print(f"✅ 粒度检查: 通过")
    
    print(f"="*70)
    print(f"\n✅ 完成！输出文件: {args.output}")


def main():
    """主函数"""
    args = parse_args()
    
    try:
        main_process(args)
    except KeyboardInterrupt:
        print("\n\n⚠️  已取消")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
